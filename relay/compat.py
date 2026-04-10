"""
relay/compat.py

Lecture de la matrice de compatibilité depuis un fichier Excel (.xlsx).

Le fichier doit contenir un triangle inférieur de scores entiers ≥ 0,
avec la diagonale marquée "X" et le triangle supérieur vide.
"""

import openpyxl


def read_compat_matrix(path: str) -> dict[tuple[str, str], int]:
    """Lit un fichier Excel et retourne le triangle inférieur canonique.

    Retourne un dict {(nom_a, nom_b): score} avec a < b selon l'ordre
    du fichier. constraints.py reconstruit la symétrie à la lecture.
    """
    print(f"Chargement compatibilités : {path}")
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Upper-left anchor is B4 (row 4, col 2); detect last row/col dynamically
    min_row, min_col = 4, 2
    max_row = max(
        (r for r in range(min_row, ws.max_row + 1) if ws.cell(r, min_col).value is not None),
        default=min_row,
    )
    max_col = max(
        (c for c in range(min_col, ws.max_column + 1) if ws.cell(min_row, c).value is not None),
        default=min_col,
    )
    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    header = [c for c in rows[0][1:] if c is not None]
    n = len(header)

    # Structural checks
    struct_errors = []
    data_rows = [r for r in rows[1:] if r[0] is not None]
    if len(data_rows) != n:
        struct_errors.append(
            f"  matrice non carrée : {len(data_rows)} lignes de données, {n} colonnes"
        )
    row_labels = [r[0] for r in data_rows]
    for i, (row_name, col_name) in enumerate(zip(row_labels, header)):
        if row_name != col_name:
            struct_errors.append(
                f"  position {i + 1} : ligne '{row_name}' ≠ colonne '{col_name}'"
            )
    if len(row_labels) > len(header):
        for name in row_labels[len(header):]:
            struct_errors.append(f"  ligne sans colonne correspondante : '{name}'")
    elif len(row_labels) < len(header):
        for name in header[len(row_labels):]:
            struct_errors.append(f"  colonne sans ligne correspondante : '{name}'")
    # Diagonal must be "X"
    for i, row_name in enumerate(r[0] for r in data_rows):
        diag_val = data_rows[i][i + 1]
        if str(diag_val).strip().upper() != "X":
            struct_errors.append(
                f"  diagonale ({row_name}, {row_name}): attendu 'X', trouvé {diag_val!r}"
            )
    # Upper triangle must be empty
    for i in range(min(len(data_rows), n)):
        row_name = data_rows[i][0]
        for j in range(i + 1, n):
            col_name = header[j]
            val = data_rows[i][j + 1]
            if val is not None:
                struct_errors.append(
                    f"  triangle supérieur ({row_name}, {col_name}): attendu vide, trouvé {val!r}"
                )
    if struct_errors:
        raise ValueError("Erreurs de structure dans la matrice :\n" + "\n".join(struct_errors))

    compat: dict[tuple[str, str], int] = {}
    errors = []
    idx = {name: i for i, name in enumerate(header)}
    for data_row in rows[1:]:
        row_name = data_row[0]
        if row_name is None:
            continue
        for col_idx, col_name in enumerate(header):
            val = data_row[col_idx + 1]
            if row_name == col_name:
                continue  # diagonale
            if val is None:
                continue  # triangle supérieur
            if isinstance(val, float) and val.is_integer():
                val = int(val)
            if not isinstance(val, int) or val < 0:
                errors.append(f"  ({row_name}, {col_name}): {val!r} n'est pas un entier naturel")
                continue
            # Stocker en clé canonique (a, b) avec idx[a] < idx[b]
            a, b = (row_name, col_name) if idx[row_name] < idx[col_name] else (col_name, row_name)
            compat[(a, b)] = val

    if errors:
        raise ValueError("Valeurs invalides dans la matrice :\n" + "\n".join(errors))

    return compat
