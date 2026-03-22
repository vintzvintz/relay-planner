"""
Lit compat_coureurs.xlsx et génère compat.py avec la matrice de compatibilité
et la fonction is_compatible().
"""

import openpyxl

XLSX_PATH = "compat_coureurs.xlsx"
OUTPUT_PATH = "compat.py"


def read_matrix(path: str) -> tuple[list[str], dict[tuple[str, str], bool]]:
    wb = openpyxl.load_workbook(path)
    ws = wb.active

    # Upper-left anchor is B4 (row 4, col 2); detect last row/col dynamically
    min_row, min_col = 4, 2
    max_row = max(
        (r for r in range(min_row, ws.max_row + 1) if ws.cell(r, min_col).value is not None),
        default=min_row,
    )
    max_col = max(
        (c for c in range(min_col, ws.max_column + 1) if ws.cell(min_row, c).value is not None),
        default=min_col,
    )
    rows = list(ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col, values_only=True))
    header = [c for c in rows[0][1:] if c is not None]
    n = len(header)

    # Structural checks
    struct_errors = []
    # 1. Square: number of data rows must equal number of header columns
    data_rows = [r for r in rows[1:] if r[0] is not None]
    if len(data_rows) != n:
        struct_errors.append(
            f"  matrice non carrée : {len(data_rows)} lignes de données, {n} colonnes"
        )
    # 1b. Row labels must match column headers (same names, same order)
    row_labels = [r[0] for r in data_rows]
    for i, (row_name, col_name) in enumerate(zip(row_labels, header)):
        if row_name != col_name:
            struct_errors.append(
                f"  position {i + 1} : ligne '{row_name}' ≠ colonne '{col_name}'"
            )
    if len(row_labels) == len(header) and row_labels != header:
        pass  # individual mismatches already reported above
    elif len(row_labels) > len(header):
        for name in row_labels[len(header):]:
            struct_errors.append(f"  ligne sans colonne correspondante : '{name}'")
    elif len(row_labels) < len(header):
        for name in header[len(row_labels):]:
            struct_errors.append(f"  colonne sans ligne correspondante : '{name}'")
    # 2. Diagonal must be "X" (case-insensitive)
    for i, row_name in enumerate(r[0] for r in data_rows):
        diag_val = data_rows[i][i + 1]  # +1 because col 0 is the row label
        if str(diag_val).strip().upper() != "X":
            struct_errors.append(
                f"  diagonale ({row_name}, {row_name}): attendu 'X', trouvé {diag_val!r}"
            )
    # 3. Upper triangle must be empty (None)
    for i in range(min(len(data_rows), n)):
        row_name = data_rows[i][0]
        for j in range(i + 1, n):
            col_name = header[j]
            val = data_rows[i][j + 1]  # +1 for row-label offset
            if val is not None:
                struct_errors.append(
                    f"  triangle supérieur ({row_name}, {col_name}): attendu vide, trouvé {val!r}"
                )
    if struct_errors:
        raise ValueError("Erreurs de structure dans la matrice :\n" + "\n".join(struct_errors))

    compat: dict[tuple[str, str], int] = {}
    errors = []
    for data_row in rows[1:]:
        row_name = data_row[0]
        if row_name is None:
            continue
        for col_idx, col_name in enumerate(header):
            val = data_row[col_idx + 1]
            if row_name == col_name:
                continue  # diagonale, ignorée
            if val is None:
                val = 0
            elif isinstance(val, float) and val.is_integer():
                val = int(val)
            if not isinstance(val, int) or val < 0:
                errors.append(f"  ({row_name}, {col_name}): {val!r} n'est pas un entier naturel")
                continue
            compat[(row_name, col_name)] = val
            compat[(col_name, row_name)] = val

    if errors:
        raise ValueError("Valeurs invalides dans la matrice :\n" + "\n".join(errors))

    return header, compat


def generate_compat_py(runners: list[str], compat: dict[tuple[str, str], int]) -> str:
    lines = [
        '# -*- coding: utf-8 -*-',
        '"""',
        "Matrice de compatibilité générée automatiquement par refresh_compat.py.",
        'Ne pas modifier manuellement — éditer compat_coureurs.xlsx puis relancer refresh_compat.py.',
        '"""',
        "",
        "# fmt: off",
        f"RUNNERS: list[str] = {runners!r}",
        "",
        "# Triangle inférieur uniquement (clé canonique : (a, b) avec a < b selon RUNNERS).",
        "# constraints.py reconstruit la symétrie à la lecture via compat_score().",
        "COMPAT_MATRIX: dict[tuple[str, str], int] = {",
    ]

    idx = {name: i for i, name in enumerate(runners)}
    for r1 in runners:
        for r2 in runners:
            if idx[r1] >= idx[r2]:
                continue  # ne stocker que (r_petit_idx, r_grand_idx)
            val = compat.get((r1, r2), 0)
            lines.append(f'    ("{r1}", "{r2}"): {val},')

    lines += [
        "}",
        "# fmt: on",
    ]

    return "\n".join(lines) + "\n"


if __name__ == "__main__":
    runners, compat = read_matrix(XLSX_PATH)
    content = generate_compat_py(runners, compat)
    with open(OUTPUT_PATH, "w", encoding="utf-8") as f:
        f.write(content)
    print(f"Généré : {OUTPUT_PATH}  ({len(runners)} coureurs, {sum(v > 0 for v in compat.values()) // 2} paires compatibles)")
